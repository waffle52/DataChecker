#!/usr/bin/env python3
# This script creates a gui for the DataChecker
import os
import sys

from PyQt5.QtWidgets import *
from PyQt5 import *
from PyQt5.QtGui import QPixmap
from PyQt5 import QtCore, uic, QtWidgets
from PyQt5.QtGui import QIcon

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import utils
from funcs import *
from datatypes import *
from File_class import *

Ui_MainWindow, QtBaseClass = uic.loadUiType("layout.ui")


class MyApp(QtWidgets.QMainWindow, Ui_MainWindow):
    File = ""

    def __init__(self):
        #initialize GUI
        self.File = ""
        QtWidgets.QMainWindow.__init__(self)
        Ui_MainWindow.__init__(self)
        self.setupUi(self)
        self.setWindowTitle("WestCal DataChecker")
        self.Check_File_Btn.clicked.connect(self.Check)
        self.Select_File_Btn.clicked.connect(self.Select)

    def get_file(self):
        return self.File

    def set_file(self, name):
        self.File = name

    def Check(self):
        self.main()

    def Select(self):
        self.openFileNameDialog()

    def openFileNameDialog(self):
        # prompt user for which file to load
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self, "QFileDialog.getOpenFileName()", "","All Files (*);;Python Files (*.py)", options=options)
        if fileName:
            self.set_file(fileName)
            self.File_Name_Lbl.setText("{}".format(os.path.basename(self.get_file())))
            print(fileName)

    def show_popup(self, text):
        # popup messages detailing errors
        msg = QMessageBox()
        msg.setText("Results:")
        msg.setStandardButtons(QMessageBox.Ignore)
        msg.setDefaultButton(QMessageBox.Ignore)
        msg.setInformativeText(text)
        x = msg.exec_()

    def main(self):
        # main function to check an excel file format
        if (self.get_file() == ""):
            print("TEST")
            return
        file_name = self.get_file()
        # check if the file passed is an excel file
        try:
            wb = load_workbook(file_name)
        except (FileNotFoundError, utils.exceptions.InvalidFileException):
            print("Cannot open. Program only designed for excel files")
            sys.exit(1)
        # variables
        ws = wb.worksheets[0]
        func_num = 0
        first_try = 0
        track = 1
        row_max = ws.max_row + 1
        column_max = ws.max_column + 1
        column_name = ws.cell(1, 1).value
        prev = column_name
        test = False

        # check if input is blank then use this otherwise set num
        text = self.Num_Input.text()
        stop = findBlank(ws, column_max, row_max) - 1
        if (isinstance(text, int) is True and text < stop and text > 0):
            stop = text

        # create Results.txt file to report any errors
        f = open("Results.txt", "w")
        f.write("WESTCAL CHECKER ERRORS:\n")
        f.close()
        f = open("Results.txt", "a")
        info = File(0, 0, f, 0)

        # loop through length of data present in the excel file passed
        for column_x in range(1, column_max):
            info.set_info('x', column_x)
            info.set_info("name", ws.cell(1, column_x).value)
            for row_y in range(2, row_max):
                info.set_info('y', row_y)
                column_name = ws.cell(1, column_x).value
                if (track == stop):
                    track = 1
                    break
                if (column_name != prev):
                    prev = column_name
                    func_num += 1
                if (func_list[func_num](ws.cell(row_y, column_x).value, info) is False):
                    self.mark(column_name, column_x, row_y, f)
                track += 1
        print("Done!")
        f.close()

        text = open("Results.txt", 'r')
        self.show_popup(text.read())
        text.close()

    def mark(self, column_name, x, y, f):
        f.write("Error with {} value: Column: {} - Row: {}\n"
                .format(column_name, x, y))


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MyApp()
    window.show()
    sys.exit(app.exec_())
